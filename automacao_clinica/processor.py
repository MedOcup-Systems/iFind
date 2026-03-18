"""
processor.py — Lógica de busca e extração de documentos PDF.
Versão final corrigida — compatível com app.py (interface Streamlit).

Correções aplicadas sobre o doc1 (versão quebrada):
  1. montar_caminho_pasta + autodetectar_meses restaurados do doc2
     (doc1 montava drive/janeiro/15; drive real usa 1.JANEIRO 2025/15.01)
  2. processar_lista: assinatura do callback corrigida para 4 args
     cb(prog, etapa, detalhe, status) — igual ao que o app.py espera
  3. Parâmetros modo_extracao, dpi_ocr, max_workers restaurados
     (app.py passa esses parâmetros; doc1 não os aceitava → TypeError)
  4. validar_planilha restaurada (app.py importa e chama essa função)
  5. filtrar_aso REMOVIDO de processar_lista
     (doc2 não tem; app.py novo não passa; era exclusividade do doc1)
  6. Cache OCR persistente mantido (melhoria de performance do doc2)
  7. Worker paralelo mantido (melhoria de performance do doc2)
"""

import fitz
import openpyxl
import unicodedata
import re
import io
import sys
import os
import threading
import hashlib
import json as _json
from pathlib import Path
from datetime import datetime, date
from concurrent.futures import ThreadPoolExecutor, as_completed

# Imports opcionais — carregados UMA VEZ no nível do módulo.
# Evita overhead de 'from rapidfuzz import fuzz' dentro de loops de milhares de páginas.
try:
    from rapidfuzz import fuzz as _fuzz
    _RAPIDFUZZ_OK = True
except ImportError:
    _fuzz = None
    _RAPIDFUZZ_OK = False

try:
    from PIL import Image as _PILImage
    import pytesseract as _pytesseract
    _OCR_OK = True
except ImportError:
    _PILImage = None
    _pytesseract = None
    _OCR_OK = False


# ---------------------------------------------------------------------------
# Cache OCR persistente
# ---------------------------------------------------------------------------

def _cache_pdf_path(caminho_pdf: Path) -> Path:
    return caminho_pdf.parent / f".ocr_{caminho_pdf.stem}.json"

def _pdf_hash(caminho_pdf: Path) -> str:
    try:
        with open(caminho_pdf, "rb") as f:
            return hashlib.md5(f.read(65536)).hexdigest()
    except Exception:
        return ""

def _cache_load(caminho_pdf: Path) -> dict:
    p = _cache_pdf_path(caminho_pdf)
    if not p.exists():
        return {}
    try:
        d = _json.loads(p.read_text(encoding="utf-8"))
        if d.get("hash") != _pdf_hash(caminho_pdf):
            p.unlink(missing_ok=True)
            return {}
        return d.get("paginas", {})
    except Exception:
        return {}

def _cache_save(caminho_pdf: Path, paginas: dict):
    try:
        _cache_pdf_path(caminho_pdf).write_text(
            _json.dumps({
                "hash": _pdf_hash(caminho_pdf),
                "gerado_em": datetime.now().isoformat(),
                "paginas": paginas,
            }, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception as e:
        print(f"[Cache] aviso ao salvar: {e}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Utilitários de texto
# ---------------------------------------------------------------------------

def normalizar_texto(texto: str) -> str:
    if not texto:
        return ""
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", texto.lower()).strip()


def nome_contem(texto_pagina: str, nome_buscado: str, threshold: float = 80.0) -> tuple[bool, float]:
    texto_norm = normalizar_texto(texto_pagina)
    nome_norm  = normalizar_texto(nome_buscado)

    palavras = nome_norm.split()
    if palavras and all(p in texto_norm for p in palavras):
        return True, 100.0

    if _RAPIDFUZZ_OK:
        palavras_texto = texto_norm.split()
        tam = len(palavras)
        melhor = 0.0
        for i in range(max(1, len(palavras_texto) - tam + 1)):
            janela = " ".join(palavras_texto[i: i + tam + 2])
            score  = _fuzz.token_set_ratio(nome_norm, janela)
            if score > melhor:
                melhor = score
        if melhor >= threshold:
            return True, float(melhor)

    return False, 0.0


# ---------------------------------------------------------------------------
# Seletor de pasta nativo
# ---------------------------------------------------------------------------

def abrir_seletor_pasta(titulo: str = "Selecionar pasta") -> str:
    import platform
    sistema = platform.system().lower()

    if sistema == "windows":
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.wm_attributes("-topmost", True)
            pasta = filedialog.askdirectory(title=titulo)
            root.destroy()
            return pasta or ""
        except Exception as e:
            print(f"Erro ao abrir seletor Windows: {e}", file=sys.stderr)
            return ""
    elif sistema == "darwin":
        try:
            import subprocess
            script = f'choose folder with prompt "{titulo}"'
            r = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
            if r.returncode == 0:
                caminho = r.stdout.strip().replace(":", "/")
                return caminho if caminho.startswith("/") else "/" + caminho
        except Exception:
            return ""
    else:
        for cmd in [["zenity", "--file-selection", "--directory", f"--title={titulo}"],
                    ["yad", "--file", "--directory", f"--title={titulo}"]]:
            try:
                import subprocess
                import shutil
                if not shutil.which(cmd[0]):
                    continue
                r = subprocess.run(cmd, capture_output=True, text=True)
                if r.returncode == 0:
                    return r.stdout.strip()
            except Exception:
                continue
    return ""


# ---------------------------------------------------------------------------
# Leitura inteligente da planilha Excel
# ---------------------------------------------------------------------------

_PALAVRAS_NOME  = {"nome", "paciente", "patient", "name", "beneficiario", "segurado", "cliente", "pessoa", "titular"}
_PALAVRAS_DATA  = {"data", "date", "dt", "dia", "procedimento", "proc", "atendimento", "consulta", "exame", "realizacao", "competencia"}
_PALAVRAS_EMAIL = {"email", "e-mail", "correio", "mail"}
_PALAVRAS_LIXO  = {"total", "subtotal", "soma", "media", "count", "grand total", "totais", "#", "n/a", "n.a."}

def _celula_para_str(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (datetime, date)):
        return str(valor)
    return str(valor).strip()

def _parece_data(valor) -> bool:
    if isinstance(valor, (datetime, date)):
        return True
    s = str(valor).strip()
    padroes = [r"^\d{1,2}/\d{1,2}(/\d{2,4})?$", r"^\d{4}-\d{2}-\d{2}", r"^\d{1,2}-\d{1,2}-\d{2,4}$"]
    return any(re.match(p, s) for p in padroes)

def _parece_email(valor) -> bool:
    s = str(valor).strip() if valor else ""
    return bool(re.match(r"[^@]+@[^@]+\.[^@]+", s))

def _parece_nome(valor) -> bool:
    s = str(valor).strip() if valor else ""
    palavras = [p for p in s.split() if re.search(r"[a-zA-ZÀ-ú]", p)]
    return len(palavras) >= 2 and not _parece_data(valor) and not _parece_email(valor)

def _linha_e_lixo(linha: tuple) -> bool:
    valores = [_celula_para_str(v).lower() for v in linha if v is not None]
    if not valores:
        return True
    for v in valores:
        if any(lixo in v for lixo in _PALAVRAS_LIXO):
            return True
    return False

def _detectar_coluna(cabecalho: list[str], palavras_chave: set, fallback: int) -> int:
    for i, h in enumerate(cabecalho):
        h_norm = normalizar_texto(h)
        if any(p in h_norm for p in palavras_chave):
            return i
    return fallback

def _score_celula_cabecalho(celula_str: str) -> int:
    s = celula_str.lower().strip()
    if not s:
        return 0
    score = 0
    if s in _PALAVRAS_NOME:  score += 4
    if s in _PALAVRAS_DATA:  score += 4
    if s in _PALAVRAS_EMAIL: score += 2
    return score

def _encontrar_linha_cabecalho(ws) -> tuple[int, bool]:
    melhor_linha, melhor_score = 1, 0
    for num_linha in range(1, min(11, ws.max_row + 1)):
        score = sum(_score_celula_cabecalho(_celula_para_str(c.value)) for c in ws[num_linha])
        if score > melhor_score:
            melhor_score, melhor_linha = score, num_linha
    return melhor_linha, (melhor_score >= 4)

def _detectar_colunas_por_conteudo(ws, primeira_linha: int) -> tuple[int, int, int]:
    num_amostras = min(primeira_linha + 15, ws.max_row + 1) - primeira_linha
    if num_amostras <= 0:
        return 0, 1, -1

    score_nome  = {}
    score_data  = {}
    score_email = {}

    for num_linha in range(primeira_linha, primeira_linha + num_amostras):
        if num_linha > ws.max_row:
            break
        for col_idx, cell in enumerate(ws[num_linha]):
            v = cell.value
            if v is None:
                continue
            if _parece_nome(v):
                score_nome[col_idx]  = score_nome.get(col_idx, 0) + 1
            if _parece_data(v):
                score_data[col_idx]  = score_data.get(col_idx, 0) + 1
            if _parece_email(v):
                score_email[col_idx] = score_email.get(col_idx, 0) + 1

    col_nome  = max(score_nome,  key=score_nome.get)  if score_nome  else 0
    col_data  = max(score_data,  key=score_data.get)  if score_data  else 1
    col_email = max(score_email, key=score_email.get) if score_email else -1

    if col_nome == col_data:
        candidatos_data = sorted(score_data, key=score_data.get, reverse=True)
        col_data = next((c for c in candidatos_data if c != col_nome), col_nome + 1)

    return col_nome, col_data, col_email

def _ler_aba(ws) -> list[dict]:
    if ws.max_row < 1:
        return []
    linha_cab, tem_cabecalho = _encontrar_linha_cabecalho(ws)
    if tem_cabecalho:
        cabecalho = [_celula_para_str(c.value) for c in ws[linha_cab]]
        col_nome  = _detectar_coluna(cabecalho, _PALAVRAS_NOME,  0)
        col_data  = _detectar_coluna(cabecalho, _PALAVRAS_DATA,  1)
        col_email = _detectar_coluna(cabecalho, _PALAVRAS_EMAIL, -1)
        primeira_linha_dados = linha_cab + 1
        nomes_cabecalho = {normalizar_texto(h) for h in cabecalho if h}
    else:
        col_nome, col_data, col_email = _detectar_colunas_por_conteudo(ws, 1)
        primeira_linha_dados = 1
        nomes_cabecalho = set()

    registros = []
    for num_linha in range(primeira_linha_dados, ws.max_row + 1):
        linha = tuple(c.value for c in ws[num_linha])
        if all(v is None for v in linha) or _linha_e_lixo(linha):
            continue
        nome = _celula_para_str(linha[col_nome])
        if not nome or normalizar_texto(nome) in nomes_cabecalho:
            continue
        data  = linha[col_data]  if col_data  < len(linha) else None
        email = _celula_para_str(linha[col_email]) if 0 <= col_email < len(linha) else ""
        registros.append({"nome": nome, "data": data, "email": email})
    return registros

def ler_planilha(caminho_excel: str) -> list[dict]:
    wb = openpyxl.load_workbook(caminho_excel, data_only=True)
    melhor_registros = []
    for ws in [wb.active] + [wb[n] for n in wb.sheetnames if wb[n] != wb.active]:
        try:
            regs = _ler_aba(ws)
            if len(regs) > len(melhor_registros):
                melhor_registros = regs
        except Exception:
            continue
    if not melhor_registros:
        raise ValueError("Planilha vazia ou ilegível.")
    return melhor_registros


def validar_planilha(registros: list[dict]) -> list[dict]:
    """Detecta duplicatas (mesmo nome + mesma data) e outros avisos."""
    avisos = []
    vistos: dict[tuple, int] = {}
    for i, reg in enumerate(registros, 1):
        nome = reg.get("nome", "").strip()
        data = str(reg.get("data", "")).strip()
        if len(nome.split()) < 2:
            avisos.append({"tipo": "nome_curto",
                           "msg": f"Nome muito curto na linha {i}: '{nome}'", "linha": i})
        chave = (normalizar_texto(nome), data)
        if chave in vistos:
            avisos.append({"tipo": "duplicata",
                           "msg": f"Duplicata: '{nome}' em {data} (linhas {vistos[chave]} e {i})",
                           "linha": i})
        else:
            vistos[chave] = i
        if not data or data in ("None", ""):
            avisos.append({"tipo": "data_invalida",
                           "msg": f"Data ausente na linha {i} ({nome})", "linha": i})
    return avisos


# ---------------------------------------------------------------------------
# Navegação e Caminhos
# ---------------------------------------------------------------------------

def extrair_dia_mes(data) -> tuple[str, str]:
    if isinstance(data, (datetime, date)):
        return data.strftime("%d"), data.strftime("%m")
    texto = str(data).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", texto):
        p = texto.split("-")
        return p[2][:2], p[1][:2]
    if "/" in texto:
        p = texto.split("/")
        return p[0].strip().zfill(2), p[1].strip().zfill(2)
    return "01", "01"


# Nomes padrão das pastas de mês (usados como fallback se autodetecção falhar)
MESES = {
    "01": "1.JANEIRO",  "02": "2.FEVEREIRO", "03": "3.MARCO",
    "04": "4.ABRIL",    "05": "5.MAIO",       "06": "6.JUNHO",
    "07": "7.JULHO",    "08": "8.AGOSTO",     "09": "9.SETEMBRO",
    "10": "10.OUTUBRO", "11": "11.NOVEMBRO",  "12": "12.DEZEMBRO",
}

_NOMES_MESES_PT = {
    "janeiro": "01", "jan": "01",
    "fevereiro": "02", "fev": "02",
    "marco": "03", "março": "03", "mar": "03",
    "abril": "04", "abr": "04",
    "maio": "05", "mai": "05",
    "junho": "06", "jun": "06",
    "julho": "07", "jul": "07",
    "agosto": "08", "ago": "08",
    "setembro": "09", "set": "09",
    "outubro": "10", "out": "10",
    "novembro": "11", "nov": "11",
    "dezembro": "12", "dez": "12",
}

_CACHE_MESES: dict[str, dict[str, str]] = {}


def autodetectar_meses(drive_raiz: str) -> dict[str, str]:
    """
    Detecta automaticamente os nomes reais das pastas de mês no drive.
    Ex: encontra '1.JANEIRO 2025', '2.FEVEREIRO 2025' etc. e mapeia para
    o número do mês ('01', '02'...).
    Cache por sessão para não varrer o disco a cada chamada.
    """
    if drive_raiz in _CACHE_MESES:
        return _CACHE_MESES[drive_raiz]

    raiz = Path(drive_raiz)
    if not raiz.exists():
        return {}

    mapa: dict[str, str] = {}
    for pasta in raiz.iterdir():
        if not pasta.is_dir():
            continue
        nome_lower = normalizar_texto(pasta.name)
        for palavra, num in _NOMES_MESES_PT.items():
            if palavra in nome_lower:
                existente = mapa.get(num)
                if existente is None:
                    mapa[num] = pasta.name
                else:
                    # Prefere a pasta com mais arquivos (mais provável ser a correta)
                    try:
                        if len(list((raiz / pasta.name).iterdir())) > len(list((raiz / existente).iterdir())):
                            mapa[num] = pasta.name
                    except Exception:
                        pass
                break

    _CACHE_MESES[drive_raiz] = mapa
    return mapa


def montar_caminho_pasta(drive_raiz: str, data) -> Path:
    """
    Constrói o caminho da pasta do dia no drive.
    Usa autodetecção dos nomes reais das pastas de mês.
    Formato da pasta do dia: DD.MM (ex: 15.01)
    """
    dia, mes_num = extrair_dia_mes(data)

    try:
        if isinstance(data, (datetime, date)):
            ano = str(data.year)
        else:
            texto = str(data).strip()
            m = re.match(r"^(\d{4})", texto)
            ano = m.group(1) if m else Path(drive_raiz).name
    except Exception:
        ano = Path(drive_raiz).name

    pasta_dia = f"{dia}.{mes_num}"

    mapa = autodetectar_meses(drive_raiz)
    if mes_num in mapa:
        return Path(drive_raiz) / mapa[mes_num] / pasta_dia

    nome_mes = MESES.get(mes_num, mes_num)
    return Path(drive_raiz) / f"{nome_mes} {ano}" / pasta_dia


# ---------------------------------------------------------------------------
# Configuração do Tesseract
# ---------------------------------------------------------------------------

def _configurar_tesseract() -> tuple[str, str]:
    import pytesseract

    tessdata_dir = ""
    lang = "por"

    cfg_path = Path(__file__).parent / "config_tesseract.py"
    if cfg_path.exists():
        try:
            exec(cfg_path.read_text(encoding="utf-8"), {"pytesseract": pytesseract})
        except Exception as e:
            print(f"[Tesseract] Aviso ao ler config_tesseract.py: {e}", file=sys.stderr)

    cmd_path = pytesseract.pytesseract.tesseract_cmd
    if cmd_path and os.path.exists(str(cmd_path)):
        pasta_bin = Path(cmd_path).parent
        for candidato in [pasta_bin / "tessdata",
                          pasta_bin.parent / "tessdata",
                          pasta_bin / "share" / "tessdata"]:
            if candidato.exists():
                tessdata_dir = str(candidato.resolve())
                break

    if not tessdata_dir:
        tessdata_dir = os.environ.get("TESSDATA_PREFIX", "")

    if tessdata_dir:
        os.environ["TESSDATA_PREFIX"] = tessdata_dir

    try:
        langs_disp = pytesseract.get_languages(config="")
        if "por" not in langs_disp:
            lang = "eng"
    except Exception:
        try:
            from PIL import Image as _I
            pytesseract.image_to_string(_I.new("RGB", (10, 10), "white"), lang="por")
        except Exception:
            lang = "eng"

    if lang == "eng":
        print("[Tesseract] Usando lang=eng (por.traineddata não disponível).", file=sys.stderr)

    return tessdata_dir, lang


_TESS_CONFIG_CACHE: dict = {}

def _get_tess_config() -> tuple[str, str]:
    if not _TESS_CONFIG_CACHE:
        tessdata_dir, lang = _configurar_tesseract()
        _TESS_CONFIG_CACHE["tessdata_dir"] = tessdata_dir
        _TESS_CONFIG_CACHE["lang"] = lang
    return _TESS_CONFIG_CACHE["tessdata_dir"], _TESS_CONFIG_CACHE["lang"]


# ---------------------------------------------------------------------------
# Modos de extração
# ---------------------------------------------------------------------------

MODO_AUTO   = "auto"
MODO_NATIVO = "nativo"
MODO_OCR    = "ocr"


# ---------------------------------------------------------------------------
# Filtro ASO (Atestado de Saúde Ocupacional)
# ---------------------------------------------------------------------------

# Marcadores que IDENTIFICAM um ASO legítimo (precisa ter ao menos _ASO_SCORE_MINIMO)
_ASO_MARCADORES_INCLUSAO = {
    "aso",
    "atestado de saude ocupacional",
    "admissional",
    "demissional",
    "retorno ao trabalho",
    "mudanca de riscos ocupacionais",
    "periodico",
    "nr-7",
    "pcmso",
    "exame clinico",
    "riscos",
    "inapto",
    "apto",
    "medico responsavel",
}

# Marcadores que EXCLUEM a página — ela não é um ASO mesmo que tenha o nome
_ASO_MARCADORES_EXCLUSAO = {
    "prontuario de pericia medica",
    "prontuario",
    "pericia medica",
    "cid:",
    "dias de afastamento",
    "data do atestado",
    "afastamento total",
    "afastamento parcial",
    "receita medica",
    "prescricao medica",
    "atestado medico",
    "laudo medico",
    "relatorio medico",
}

# Pontuação mínima de marcadores de inclusão para aceitar a página como ASO
_ASO_SCORE_MINIMO = 2


def _pagina_e_aso(texto: str) -> bool:
    """
    Verifica se uma página é um ASO (Atestado de Saúde Ocupacional).

    Lógica:
      1. Normaliza o texto (remove acentos, minúsculas)
      2. Verifica se algum marcador de EXCLUSÃO está presente → não é ASO
      3. Conta marcadores de INCLUSÃO presentes
      4. É ASO se score >= _ASO_SCORE_MINIMO e nenhuma exclusão

    Retorna True se a página é ASO, False caso contrário.
    """
    texto_norm = normalizar_texto(texto)

    # Exclusões têm prioridade (mais rápido e mais seguro)
    for excluir in _ASO_MARCADORES_EXCLUSAO:
        if excluir in texto_norm:
            return False

    score = sum(1 for inc in _ASO_MARCADORES_INCLUSAO if inc in texto_norm)
    return score >= _ASO_SCORE_MINIMO


def extrair_texto_pagina(pagina: fitz.Page, forcar_ocr: bool = False, dpi: int = 150) -> str:
    """
    Extrai texto de uma página PDF.
    - forcar_ocr=False: retorna texto nativo (get_text). Se vazio, tenta OCR.
    - forcar_ocr=True : pula get_text e vai direto ao OCR.
    PIL e pytesseract são usados das variáveis de módulo (_PILImage, _pytesseract)
    para evitar import a cada chamada.
    """
    if not forcar_ocr:
        texto = pagina.get_text()
        if len(texto.strip()) >= 50:
            return texto
        # Texto nativo insuficiente — cai no OCR abaixo
    else:
        texto = ""

    if not _OCR_OK:
        return texto  # OCR não disponível, retorna o que tiver

    try:
        _, lang = _get_tess_config()
        pix = pagina.get_pixmap(dpi=dpi)
        img = _PILImage.open(io.BytesIO(pix.tobytes("png")))
        texto_ocr = _pytesseract.image_to_string(img, lang=lang)
        return texto + "\n" + texto_ocr
    except Exception as e:
        print(f"[OCR] Erro na página: {e}", file=sys.stderr)
        return texto


def buscar_nome_em_pdf(
    caminho_pdf: Path,
    nome_buscado: str,
    threshold: float = 80.0,
    callback_pagina=None,
    modo: str = MODO_AUTO,
    dpi: int = 150,
    filtrar_aso: bool = False,
) -> list[tuple[int, float]]:
    """
    Busca o nome em todas as páginas do PDF. Retorna lista de (num_pagina, score).

    Parâmetro filtrar_aso:
        Se True, ignora páginas que não sejam ASO (Atestado de Saúde Ocupacional).
        Prontuários, atestados médicos comuns, receitas etc. são pulados mesmo
        que contenham o nome do paciente.
        Default: False (busca em qualquer tipo de documento).
    """
    encontradas = []
    doc = fitz.open(str(caminho_pdf))
    total_pags = len(doc)

    for num, pagina in enumerate(doc):

        if modo == MODO_NATIVO:
            if callback_pagina:
                callback_pagina(num, total_pags, "texto nativo")
            texto = extrair_texto_pagina(pagina, forcar_ocr=False, dpi=dpi)
            if filtrar_aso and not _pagina_e_aso(texto):
                continue
            achou, score = nome_contem(texto, nome_buscado, threshold)

        elif modo == MODO_OCR:
            if callback_pagina:
                callback_pagina(num, total_pags, "pré-filtro")
            texto_rapido = pagina.get_text()
            achou_rapido, score_rapido = nome_contem(texto_rapido, nome_buscado, threshold)
            if achou_rapido:
                achou, score = achou_rapido, score_rapido
            else:
                if callback_pagina:
                    callback_pagina(num, total_pags, "OCR")
                texto = extrair_texto_pagina(pagina, forcar_ocr=True, dpi=dpi)
                if filtrar_aso and not _pagina_e_aso(texto):
                    continue
                achou, score = nome_contem(texto, nome_buscado, threshold)

        else:  # AUTO
            if callback_pagina:
                callback_pagina(num, total_pags, "texto nativo")
            texto = extrair_texto_pagina(pagina, forcar_ocr=False, dpi=dpi)
            achou, score = nome_contem(texto, nome_buscado, threshold)

            if not achou or score < threshold:
                if callback_pagina:
                    callback_pagina(num, total_pags, "OCR")
                texto_ocr = extrair_texto_pagina(pagina, forcar_ocr=True, dpi=dpi)
                achou_p, score_p = nome_contem(texto_ocr, nome_buscado, threshold)
                if score_p > score:
                    achou, score = achou_p, score_p
                    texto = texto_ocr  # usa o texto mais rico para checar ASO

            if filtrar_aso and not _pagina_e_aso(texto):
                continue

        if achou:
            encontradas.append((num, score))
            if score >= threshold:
                break  # score suficiente — para de varrer o PDF

    doc.close()
    return encontradas


def sanitizar_nome_arquivo(nome: str) -> str:
    sem_acento = unicodedata.normalize("NFD", nome)
    sem_acento = "".join(c for c in sem_acento if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", "_", re.sub(r"[^\w\s-]", "", sem_acento).strip())


def extrair_e_salvar_paginas(
    caminho_pdf: Path,
    paginas_scores: list[tuple[int, float]],
    pasta_destino: Path,
    nome_paciente: str,
) -> Path:
    pasta_destino.mkdir(parents=True, exist_ok=True)
    nome_base = sanitizar_nome_arquivo(nome_paciente)
    saida = pasta_destino / f"{nome_base}.pdf"
    # Evita sobrescrever se já existir
    contador = 2
    while saida.exists():
        saida = pasta_destino / f"{nome_base}_{contador}.pdf"
        contador += 1
    doc_orig = fitz.open(str(caminho_pdf))
    doc_novo = fitz.open()
    for idx, _ in sorted(paginas_scores):
        doc_novo.insert_pdf(doc_orig, from_page=idx, to_page=idx)
    doc_novo.save(str(saida))
    doc_novo.close()
    doc_orig.close()
    return saida


# ---------------------------------------------------------------------------
# Worker de PDF — função de MÓDULO (fora de qualquer loop)
# Sem closure, sem captura de variável de loop — seguro para ThreadPoolExecutor
# ---------------------------------------------------------------------------

def _pdf_e_digital(doc: fitz.Document, amostras: int = 5) -> bool:
    """
    Verifica rapidamente se um PDF é digital (texto nativo) ou escaneado.
    Amostra até `amostras` páginas distribuídas pelo documento.
    Custo: apenas get_text() em poucas páginas, sem nenhum OCR.
    """
    total = len(doc)
    if total == 0:
        return False
    indices = [int(i * total / min(amostras, total)) for i in range(min(amostras, total))]
    com_texto = sum(1 for i in indices if len(doc[i].get_text().strip()) >= 50)
    return com_texto >= max(1, len(indices) // 2)


def _worker_pdf(
    j: int,
    pdf: Path,
    nomes: list[str],
    threshold: float,
    modo: str,
    dpi: int,
    n_pdfs: int,
    filtrar_aso: bool = False,
) -> tuple[int, Path, dict[str, tuple[list, float, int]], list[str]]:
    """
    Processa um PDF em thread secundária.
    NÃO chama callback diretamente (Streamlit proíbe acesso à UI de threads).
    Retorna (j, pdf, resultados, logs).

    Otimizações:
      - Detecção antecipada de PDF digital vs escaneado no modo AUTO
        (amostra 5 páginas e decide o modo uma vez, evita tentativa nativa
        inútil em cada página de um PDF escaneado)
      - Early-exit por página: para de processar o PDF assim que todos
        os nomes buscados já foram encontrados com score >= threshold
      - get_text() direto no modo nativo (sem passar por extrair_texto_pagina)
      - Imports PIL/pytesseract/rapidfuzz resolvidos no nível do módulo
    """
    logs: list[str] = []

    cache = _cache_load(pdf)
    dirty = False
    textos: dict[int, str] = {}
    textos_descartados: dict[int, str] = {}

    try:
        doc = fitz.open(str(pdf))
    except Exception as e:
        logs.append(f"ERRO ao abrir {pdf.name}: {e}")
        return j, pdf, {}, logs

    total_pags = len(doc)

    # Detecção antecipada: no modo AUTO, decide uma vez se o PDF é digital ou scan
    modo_efetivo = modo
    if modo == MODO_AUTO:
        eh_digital = _pdf_e_digital(doc)
        modo_efetivo = MODO_NATIVO if eh_digital else MODO_OCR
        logs.append(
            f"PDF {j+1}/{n_pdfs}  tipo: {'digital' if eh_digital else 'escaneado'}"
            f" → modo {modo_efetivo}  {pdf.name}"
        )

    # Conjunto de nomes ainda não encontrados com score suficiente
    # Usado para early-exit: quando vazio, para o loop de páginas
    nomes_pendentes = set(nomes)

    for num, pagina in enumerate(doc):

        # Early-exit: todos os nomes já achados com score suficiente
        if not nomes_pendentes:
            logs.append(
                f"PDF {j+1}/{n_pdfs}  early-exit na pág {num+1} "
                f"(todos os nomes encontrados)  {pdf.name}"
            )
            break

        chave = f"{num}_{dpi}_{modo_efetivo}"

        if chave in cache:
            texto = cache[chave]
            logs.append(f"PDF {j+1}/{n_pdfs}  pág {num+1}/{total_pags} (cache)  {pdf.name}")
        else:
            # Extração conforme modo efetivo — sem get_text() redundante
            if modo_efetivo == MODO_NATIVO:
                logs.append(f"PDF {j+1}/{n_pdfs}  pág {num+1}/{total_pags} (nativo)  {pdf.name}")
                texto = pagina.get_text()
            else:  # MODO_OCR
                logs.append(f"PDF {j+1}/{n_pdfs}  pág {num+1}/{total_pags} (OCR)  {pdf.name}")
                texto = extrair_texto_pagina(pagina, forcar_ocr=True, dpi=dpi)

            if texto.strip():
                cache[chave] = texto
                dirty = True

        # Classifica a página (ASO ou descartada)
        if filtrar_aso and not _pagina_e_aso(texto):
            textos_descartados[num] = texto
        else:
            textos[num] = texto
            # Atualiza early-exit: remove nomes já encontrados com score suficiente
            for nome in list(nomes_pendentes):
                achou, score = nome_contem(texto, nome, threshold)
                if achou and score >= threshold:
                    nomes_pendentes.discard(nome)

    doc.close()

    if dirty:
        _cache_save(pdf, cache)

    # Varredura final — coleta todas as páginas com match para cada nome
    resultados: dict[str, tuple[list, float, int]] = {}
    for nome in nomes:
        encontradas: list[tuple[int, float]] = []
        melhor_score = 0.0
        achou_em_descartada = 0

        for num, texto in textos.items():
            achou, score = nome_contem(texto, nome, threshold)
            if score > melhor_score:
                melhor_score = score
            if achou:
                encontradas.append((num, score))
                if score >= 100.0:
                    break

        if filtrar_aso:
            for num, texto in textos_descartados.items():
                achou, _ = nome_contem(texto, nome, threshold)
                if achou:
                    achou_em_descartada += 1

        if encontradas:
            encontradas.sort(key=lambda x: x[1], reverse=True)
        resultados[nome] = (encontradas, melhor_score, achou_em_descartada)

    logs.append(f"PDF {j+1}/{n_pdfs} concluído  {pdf.name}")
    return j, pdf, resultados, logs

    logs.append(f"PDF {j+1}/{n_pdfs} concluído  {pdf.name}")
    return j, pdf, resultados, logs


# ---------------------------------------------------------------------------
# Função Principal
# ---------------------------------------------------------------------------

def processar_lista(
    caminho_excel: str,
    drive_raiz: str,
    pasta_destino: str,
    threshold_fuzzy: float = 80.0,
    callback=None,
    modo_extracao: str = MODO_AUTO,
    dpi_ocr: int = 150,
    max_workers: int = 4,
    filtrar_aso: bool = False,
) -> list[dict]:
    """
    Processa todos os registros da planilha.

    Parâmetros:
        caminho_excel   : caminho para o .xlsx
        drive_raiz      : raiz do drive com as pastas dos meses
        pasta_destino   : onde salvar os PDFs extraídos
        threshold_fuzzy : pontuação mínima para busca fuzzy (0-100)
        callback        : função(prog: float, etapa: str, detalhe: str, status: str)
                          status: "info" | "ok" | "erro" | "aviso" | "pdf_salvo"
        modo_extracao   : MODO_AUTO | MODO_NATIVO | MODO_OCR
        dpi_ocr         : resolução para OCR (150 / 200 / 300)
        max_workers     : threads paralelas para PDFs
        filtrar_aso     : se True, só extrai páginas que sejam ASO.
                          Ignora prontuários, atestados médicos comuns, receitas etc.
                          Default: False (extrai qualquer página com o nome)

    Retorna lista de dicts: nome, data, email, encontrado, arquivo, erro, score_fuzzy
    """
    def _cb(prog: float, etapa: str, detalhe: str = "", status: str = "info"):
        if callback:
            callback(prog, etapa, detalhe, status)

    # Limpa cache de detecção de meses para garantir leitura fresca do drive
    _CACHE_MESES.clear()

    _cb(0.0, "Lendo planilha", "Aguarde...")
    registros = ler_planilha(caminho_excel)
    total = len(registros)
    _cb(0.0, "Planilha lida", f"{total} registros encontrados", "ok")

    # Agrupa registros por pasta (mesma data = mesma pasta = mesmo grupo)
    grupos: dict[str, list[dict]] = {}
    for reg in registros:
        try:
            chave = str(montar_caminho_pasta(drive_raiz, reg["data"]))
        except Exception:
            chave = "__erro__"
        grupos.setdefault(chave, []).append(reg)

    n_grupos = len(grupos)
    _cb(0.0, "Agrupamento concluído",
        f"{n_grupos} pasta(s) distintas para {total} registro(s)", "ok")

    # Mapa de resultados indexado por nome
    resultados_map: dict[str, dict] = {}
    for reg in registros:
        resultados_map[reg["nome"]] = {
            "nome": reg["nome"], "data": str(reg["data"]),
            "email": reg.get("email", ""),
            "encontrado": False, "arquivo": "", "erro": "", "score_fuzzy": 0.0,
        }

    # Retomada automática: pula nomes cujo PDF já existe no destino
    destino_path = Path(pasta_destino)
    pulados = 0
    for reg in registros:
        nome = reg["nome"]
        arq = destino_path / (sanitizar_nome_arquivo(nome) + ".pdf")
        if arq.exists():
            resultados_map[nome].update({
                "encontrado": True, "arquivo": str(arq), "score_fuzzy": 100.0,
            })
            pulados += 1

    if pulados > 0:
        _cb(0.0, "Retomando execução",
            f"{pulados} já encontrado(s) na pasta de destino — pulando", "ok")

    grupos_feitos = 0

    for chave, regs_grupo_orig in grupos.items():
        grupos_feitos += 1
        prog_base = (grupos_feitos - 1) / n_grupos
        prog_fim  = grupos_feitos       / n_grupos

        # Remove nomes já encontrados (retomada)
        regs_grupo  = [r for r in regs_grupo_orig
                       if not resultados_map[r["nome"]]["encontrado"]]
        nomes_grupo = [r["nome"] for r in regs_grupo]
        n_nomes     = len(nomes_grupo)

        if n_nomes == 0:
            _cb(prog_fim, "Grupo já concluído", Path(chave).name, "ok")
            continue

        if chave == "__erro__":
            for r in regs_grupo:
                resultados_map[r["nome"]]["erro"] = "Erro ao montar caminho da pasta."
            _cb(prog_fim, "Erro de caminho", "", "erro")
            continue

        pasta = Path(chave)
        _cb(prog_base, "Verificando pasta",
            f"{pasta.name}  —  {n_nomes} paciente(s)")

        if not pasta.exists():
            for r in regs_grupo:
                resultados_map[r["nome"]]["erro"] = f"Pasta não encontrada: {pasta}"
            _cb(prog_fim, "Pasta não encontrada", str(pasta), "erro")
            continue

        pdfs   = sorted(pasta.glob("*.pdf"))
        n_pdfs = len(pdfs)

        if n_pdfs == 0:
            for r in regs_grupo:
                resultados_map[r["nome"]]["erro"] = "Nenhum PDF na pasta."
            _cb(prog_fim, "Sem PDFs", str(pasta), "aviso")
            continue

        _cb(prog_base, "Pasta localizada",
            f"{n_pdfs} PDF(s)  —  buscando {n_nomes} nome(s) de uma vez", "ok")

        # melhor[nome] = (pdf, [(pag, score)], score_max, n_paginas_descartadas_aso)
        melhor: dict[str, tuple] = {}
        lock   = threading.Lock()
        parar  = threading.Event()

        workers = min(max_workers, n_pdfs)
        with ThreadPoolExecutor(max_workers=workers) as ex:

            futuros = {
                ex.submit(
                    _worker_pdf,
                    j, pdf,
                    list(nomes_grupo),
                    threshold_fuzzy,
                    modo_extracao,
                    dpi_ocr,
                    n_pdfs,
                    filtrar_aso,
                ): (j, pdf)
                for j, pdf in enumerate(pdfs)
            }

            for fut in as_completed(futuros):
                if parar.is_set():
                    break

                try:
                    j, pdf, res, logs = fut.result()
                except Exception as e:
                    import traceback as _tb
                    _cb(prog_base, "Erro no worker",
                        _tb.format_exc().strip().splitlines()[-1], "erro")
                    continue

                # Exibe logs do worker na thread principal (seguro para Streamlit)
                prog_pdf = prog_base + (prog_fim - prog_base) * ((j + 1) / n_pdfs)
                for linha in logs:
                    _cb(prog_pdf, linha, "", "info")

                # Atualiza melhor resultado por nome
                with lock:
                    for nome, (pags, score, em_descartada) in res.items():
                        sc_atual, desc_atual = melhor.get(nome, (None, [], 0.0, 0))[2:4]
                        if score > sc_atual:
                            melhor[nome] = (pdf, pags, score, em_descartada)
                        elif em_descartada > desc_atual:
                            # Mesmo sem melhorar score, acumula contagem de descartadas
                            entrada = melhor.get(nome, (None, [], 0.0, 0))
                            melhor[nome] = (entrada[0], entrada[1], entrada[2],
                                            entrada[3] + em_descartada)

                    todos_ok = all(
                        melhor.get(n, (None, [], 0.0, 0))[2] >= threshold_fuzzy
                        for n in nomes_grupo
                    )

                if todos_ok:
                    parar.set()
                    for f in futuros:
                        f.cancel()
                    break

        # Salva resultados do grupo
        enc_grupo = 0
        for reg in regs_grupo:
            nome = reg["nome"]
            entrada = melhor.get(nome, (None, [], 0.0, 0))
            pdf_src, pags, score, em_descartada = entrada

            if pags:
                try:
                    saida = extrair_e_salvar_paginas(
                        pdf_src, pags, Path(pasta_destino), nome
                    )
                    resultados_map[nome].update({
                        "encontrado": True,
                        "arquivo": str(saida),
                        "score_fuzzy": score,
                    })
                    _cb(prog_fim, "PDF salvo",
                        f"{nome}  score {score:.0f}%", "pdf_salvo")
                    enc_grupo += 1
                except Exception as e:
                    resultados_map[nome]["erro"] = f"Erro ao salvar: {e}"
                    _cb(prog_fim, "Erro ao salvar", str(e), "erro")
            else:
                if filtrar_aso and em_descartada > 0:
                    # Nome encontrado, mas só em páginas que não são ASO
                    msg_erro = (
                        f"Nome encontrado em {em_descartada} página(s), "
                        f"mas nenhuma delas é ASO — são prontuários ou outros documentos. "
                        f"Desative o filtro ASO para extrair mesmo assim."
                    )
                    resultados_map[nome]["erro"] = msg_erro
                    _cb(prog_fim, "Filtrado — não é ASO",
                        f"{nome}  —  {em_descartada} pág(s) descartada(s)", "aviso")
                else:
                    detalhe = f"Buscado em {n_pdfs} PDF(s)"
                    if score > 0:
                        detalhe += f" — melhor score: {score:.0f}% (threshold: {threshold_fuzzy:.0f}%)"
                    resultados_map[nome]["erro"] = "Nome não localizado nos PDFs da pasta."
                    _cb(prog_fim, "Não encontrado", detalhe, "aviso")

        _cb(prog_fim, "Pasta concluída",
            f"{pasta.name}  —  {enc_grupo}/{n_nomes} encontrados",
            "ok" if enc_grupo == n_nomes else "aviso")

    resultados = [resultados_map[r["nome"]] for r in registros]
    enc_total  = sum(r["encontrado"] for r in resultados)
    _cb(1.0, "Concluído", f"{enc_total}/{total} encontrados", "ok")
    return resultados